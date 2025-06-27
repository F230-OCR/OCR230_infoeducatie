import os
import pandas as pd
from datetime import datetime
from pathlib import Path
import re

class ExcelManager:
    def __init__(self, output_folder):
        self.output_folder = output_folder
        self.excel_file_path = os.path.join(output_folder, "Date_Persoane_OCR.xlsx")
        self.data_list = []
    
    def extract_data_from_txt(self, txt_file_path):
        """Extrage datele dintr-un fișier .txt care a fost generat de process.py"""
        try:
            with open(txt_file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            # Calculăm calea relativă față de folderul de output
            relative_path = os.path.relpath(txt_file_path, self.output_folder)
            # Eliminăm extensia .txt din calea relativă
            relative_path_no_ext = os.path.splitext(relative_path)[0]
            folder_name = os.path.dirname(relative_path) if os.path.dirname(relative_path) else "Root"
            
            # Inițializăm dicționarul cu date
            data = {
                'Nume_Fisier': os.path.splitext(os.path.basename(txt_file_path))[0],
                'Cale_Fisier': relative_path_no_ext,
                'Folder_Sursa': folder_name,
                'Data_Procesare': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'Text_Complet': content.strip()
            }
            
            # Parsăm conținutul fișierului text conform structurii din process.py
            # Structura: nume\ninitiala_tatalui\nprenume\ncnp_total\nadresa\nphone\nemail\ndoiani
            lines = content.strip().split('\n')
            
            print(f"Debug: Procesare fișier {txt_file_path}")
            print(f"Debug: Linii găsite: {lines}")
            
            # Extragem datele conform structurii exacte din process.py
            data['Nume'] = lines[0].strip() if len(lines) > 0 else ''
            data['Initiala_Tatalui'] = lines[1].strip() if len(lines) > 1 else ''
            data['Prenume'] = lines[2].strip() if len(lines) > 2 else ''
            
            # CNP - ne asigurăm că este tratat ca string pentru a evita notația științifică
            cnp_raw = lines[3].strip() if len(lines) > 3 else ''
            # Curățăm CNP-ul și ne asigurăm că este string
            data['CNP'] = str(cnp_raw).strip() if cnp_raw else ''
            
            data['Adresa'] = lines[4].strip() if len(lines) > 4 else ''
            data['2_Ani'] = lines[7].strip() if len(lines) > 7 else ''  # doiani este pe poziția 7 (index)
            
            # Informații suplimentare (incluse în Excel)
            # Telefon - păstrăm întotdeauna ca string pentru a conserva zero-urile de la început
            telefon_raw = lines[5].strip() if len(lines) > 5 else ''
            if telefon_raw:
                # Eliminăm doar .0 de la sfârșit dacă există, dar păstrăm zero-urile de la început
                telefon_clean = str(telefon_raw).strip()
                if telefon_clean.endswith('.0'):
                    telefon_clean = telefon_clean[:-2]  # Eliminăm doar .0 de la sfârșit
                data['Telefon'] = telefon_clean
            else:
                data['Telefon'] = ''
            
            data['Email'] = lines[6].strip() if len(lines) > 6 else ''
            
            # ANAF de care aparțin - folosim folder_localitate_mic (ultimul folder din ierarhie)
            data['ANAF_Apartin'] = self._get_folder_localitate_mic(folder_name)
            
            print(f"Debug: Rezultat final -> Nume: '{data['Nume']}', Inițiala: '{data['Initiala_Tatalui']}', Prenume: '{data['Prenume']}', CNP: '{data['CNP']}', Telefon: '{data['Telefon']}', Email: '{data['Email']}', 2 Ani: '{data['2_Ani']}', ANAF: '{data['ANAF_Apartin']}'")
            
            return data
            
        except Exception as e:
            print(f"Eroare la procesarea fișierului {txt_file_path}: {e}")
            return None
    
    def _get_folder_localitate_mic(self, folder_name):
        """Extrage folder_localitate_mic din calea folderului"""
        if not folder_name or folder_name == "Root":
            return "NEDETERMINAT"
        
        # Împărțim calea în părți
        parts = folder_name.split(os.sep)
        
        # Ultima parte din cale este folder_localitate_mic
        if len(parts) >= 1:
            folder_localitate_mic = parts[-1].strip()
            return folder_localitate_mic if folder_localitate_mic else "NEDETERMINAT"
        
        return "NEDETERMINAT"
    
    def _separate_name_parts(self, nume_complet):
        """Separă numele complet în nume, inițiala tatălui și prenume"""
        if not nume_complet:
            return {'nume': '', 'initiala_tatalui': '', 'prenume': ''}
            
        # Curățăm textul
        nume_complet = nume_complet.strip()
        parts = nume_complet.split()
        
        print(f"Debug: Separare nume pentru: '{nume_complet}' -> {parts}")
        
        if len(parts) == 0:
            return {'nume': '', 'initiala_tatalui': '', 'prenume': ''}
        elif len(parts) == 1:
            # Doar un cuvânt - probabil numele
            result = {'nume': parts[0], 'initiala_tatalui': '', 'prenume': ''}
            print(f"Debug: Un singur cuvânt -> {result}")
            return result
        elif len(parts) == 2:
            # Două cuvinte - verificăm dacă al doilea pare inițială
            if len(parts[1]) <= 2 and parts[1].isupper():
                # Al doilea pare inițială
                result = {'nume': parts[0], 'initiala_tatalui': parts[1], 'prenume': ''}
                print(f"Debug: Două cuvinte (cu inițială) -> {result}")
                return result
            else:
                # Nume și prenume
                result = {'nume': parts[0], 'initiala_tatalui': '', 'prenume': parts[1]}
                print(f"Debug: Două cuvinte (nume + prenume) -> {result}")
                return result
        elif len(parts) == 3:
            # Trei cuvinte - căutăm inițiala
            # Verificăm dacă al doilea cuvânt pare să fie o inițială
            if len(parts[1]) <= 2 and (parts[1].isupper() or parts[1].endswith('.')):
                result = {'nume': parts[0], 'initiala_tatalui': parts[1], 'prenume': parts[2]}
                print(f"Debug: Trei cuvinte (cu inițială în mijloc) -> {result}")
                return result
            # Verificăm dacă al treilea cuvânt pare să fie o inițială
            elif len(parts[2]) <= 2 and (parts[2].isupper() or parts[2].endswith('.')):
                result = {'nume': parts[0], 'initiala_tatalui': parts[2], 'prenume': parts[1]}
                print(f"Debug: Trei cuvinte (cu inițială la sfârșit) -> {result}")
                return result
            else:
                # Probabil nume compus sau prenume compus
                result = {'nume': parts[0], 'initiala_tatalui': '', 'prenume': ' '.join(parts[1:])}
                print(f"Debug: Trei cuvinte (fără inițială) -> {result}")
                return result
        else:
            # Mai mult de trei cuvinte
            # Căutăm o inițială (un-două caractere, de preferință mari)
            for i in range(1, len(parts)):
                if len(parts[i]) <= 2 and (parts[i].isupper() or parts[i].endswith('.')):
                    result = {
                        'nume': ' '.join(parts[:i]),
                        'initiala_tatalui': parts[i],
                        'prenume': ' '.join(parts[i+1:])
                    }
                    print(f"Debug: Multe cuvinte (cu inițială la poziția {i}) -> {result}")
                    return result
            
            # Dacă nu găsim inițială, împărțim în nume și prenume
            result = {
                'nume': parts[0],
                'initiala_tatalui': '',
                'prenume': ' '.join(parts[1:])
            }
            print(f"Debug: Multe cuvinte (fără inițială) -> {result}")
            return result
    
    def add_person_data(self, txt_file_path):
        """Adaugă datele unei persoane în lista pentru Excel"""
        data = self.extract_data_from_txt(txt_file_path)
        if data:
            self.data_list.append(data)
            return True
        return False
    
    def create_excel_file(self):
        """Creează fișierul Excel cu toate datele"""
        try:
            if not self.data_list:
                print("Nu există date pentru a crea fișierul Excel.")
                return False
            
            # Creăm DataFrame-ul
            df = pd.DataFrame(self.data_list)
            
            # Ne asigurăm că CNP-ul și Telefonul sunt tratate ca string pentru a evita formatarea automată
            if 'CNP' in df.columns:
                df['CNP'] = df['CNP'].astype(str)
            if 'Telefon' in df.columns:
                df['Telefon'] = df['Telefon'].astype(str)
            
            # Reordonăm coloanele pentru a respecta ordinea cerută
            preferred_columns = [
                'Nume', 'Initiala_Tatalui', 'Prenume', 'CNP', 'Adresa', 'ANAF_Apartin', 'Telefon', 'Email', '2_Ani'
            ]
            
            # Adăugăm coloanele suplimentare la sfârșit pentru referință
            additional_columns = [
                'Nume_Fisier', 'Folder_Sursa', 'Cale_Fisier', 'Data_Procesare', 'Text_Complet'
            ]
            
            # Reordonăm coloanele existente
            existing_main_columns = [col for col in preferred_columns if col in df.columns]
            existing_additional_columns = [col for col in additional_columns if col in df.columns]
            other_columns = [col for col in df.columns if col not in preferred_columns + additional_columns]
            
            final_columns = existing_main_columns + existing_additional_columns + other_columns
            
            df = df[final_columns]
            
            # Salvăm în Excel cu formatare
            with pd.ExcelWriter(self.excel_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Date_Persoane', index=False)
                
                # Obținem worksheet-ul pentru formatare
                worksheet = writer.sheets['Date_Persoane']
                
                # Formatăm coloanele CNP și Telefon ca text pentru a evita formatarea automată
                from openpyxl.styles import NamedStyle
                from openpyxl.utils import get_column_letter
                
                # Găsim coloanele CNP și Telefon
                columns_to_format = {'CNP': None, 'Telefon': None}
                for idx, col_name in enumerate(df.columns, 1):
                    if col_name in columns_to_format:
                        columns_to_format[col_name] = idx
                        
                print(f"Debug: Coloane de formatat găsite: {columns_to_format}")
                
                # Formatăm coloanele găsite
                for col_name, col_index in columns_to_format.items():
                    if col_index:
                        col_letter = get_column_letter(col_index)
                        # Formatăm întreaga coloană ca text (inclusiv header-ul)
                        for row in range(1, len(df) + 2):  # De la header (1) până la ultimul rând
                            cell = worksheet[f'{col_letter}{row}']
                            cell.number_format = '@'  # Format text
                            
                            # Pentru datele din rândurile de conținut (nu header)
                            if row > 1 and cell.value is not None:
                                # Convertim totul la string, indiferent de conținut
                                original_value = str(cell.value).strip()
                                
                                # Pentru CNP, dacă este în notație științifică, îl convertim
                                if col_name == 'CNP' and ('E+' in original_value or 'e+' in original_value or '.' in original_value):
                                    try:
                                        # Încercăm să convertim din notație științifică
                                        cell.value = str(int(float(original_value)))
                                    except:
                                        cell.value = original_value
                                # Pentru telefon, eliminăm .0 dacă există, dar păstrăm zero-urile de la început  
                                elif col_name == 'Telefon' and original_value.endswith('.0'):
                                    # Eliminăm doar .0 de la sfârșit, păstrând zero-urile de la început
                                    cell.value = original_value[:-2]
                                else:
                                    # Pentru toate celelalte, îl păstrăm ca string
                                    cell.value = original_value
                
                # Ajustăm lățimea coloanelor
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Maxim 50 caractere
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Fișierul Excel a fost creat cu succes: {self.excel_file_path}")
            print(f"Au fost procesate {len(self.data_list)} persoane.")
            return True
            
        except Exception as e:
            print(f"Eroare la crearea fișierului Excel: {e}")
            return False
    
    def process_all_txt_files(self, folder_path):
        """Procesează toate fișierele .txt dintr-un folder și toate subfolderele sale"""
        txt_files_found = []
        
        # Căutăm recursiv în toate folderele și subfolderele
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith('.txt'):
                    txt_path = os.path.join(root, file)
                    txt_files_found.append(txt_path)
        
        if not txt_files_found:
            print("Nu au fost găsite fișiere .txt în folderul specificat și subfolderele sale.")
            return False
        
        print(f"Au fost găsite {len(txt_files_found)} fișiere .txt în folderul și subfolderele sale:")
        for txt_file in txt_files_found:
            print(f"  - {txt_file}")
            self.add_person_data(txt_file)
        
        return self.create_excel_file()

    def _determine_anaf(self, adresa, folder_name):
        """Determină ANAF-ul pe baza adresei sau folder-ului"""
        if not adresa and folder_name == "Root":
            return "NEDETERMINAT"
        
        # Dacă avem informații din folder
        if folder_name and folder_name != "Root":
            # Încearcă să extragi informații ANAF din numele folderului
            folder_lower = folder_name.lower()
            if "anaf" in folder_lower:
                return folder_name
        
        # Dacă avem adresa, încearcă să determinisectorul/județul
        if adresa:
            adresa_lower = adresa.lower()
            
            # București - determinare sector
            if "bucuresti" in adresa_lower or "sector" in adresa_lower:
                for i in range(1, 7):
                    if f"sector {i}" in adresa_lower or f"sectorul {i}" in adresa_lower:
                        return f"ANAF SECTOR {i}"
                return "ANAF BUCURESTI"
            
            # Județe comune
            judete_anaf = {
                'cluj': 'ANAF CLUJ',
                'timis': 'ANAF TIMIS',
                'constanta': 'ANAF CONSTANTA',
                'iasi': 'ANAF IASI',
                'brasov': 'ANAF BRASOV',
                'galati': 'ANAF GALATI',
                'dolj': 'ANAF DOLJ',
                'arad': 'ANAF ARAD',
                'sibiu': 'ANAF SIBIU',
                'bacau': 'ANAF BACAU',
                'prahova': 'ANAF PRAHOVA',
                'maramures': 'ANAF MARAMURES',
                'bihor': 'ANAF BIHOR',
                'mures': 'ANAF MURES',
                'suceava': 'ANAF SUCEAVA'
            }
            
            for judet, anaf in judete_anaf.items():
                if judet in adresa_lower:
                    return anaf
        
        return "NEDETERMINAT"
    
    def add_single_record_to_excel(self, txt_file_path):
        """Adaugă o singură înregistrare direct în fișierul Excel existent"""
        try:
            # Extragem datele din fișierul .txt
            data = self.extract_data_from_txt(txt_file_path)
            if not data:
                return False
            
            # Verificăm dacă fișierul Excel există
            if os.path.exists(self.excel_file_path):
                # Citim Excel-ul existent
                try:
                    df_existing = pd.read_excel(self.excel_file_path, sheet_name='Date_Persoane')
                    
                    # Verificăm dacă această înregistrare există deja (pe baza căii fișierului)
                    if 'Cale_Fisier' in df_existing.columns:
                        relative_path_no_ext = os.path.splitext(os.path.relpath(txt_file_path, self.output_folder))[0]
                        if relative_path_no_ext in df_existing['Cale_Fisier'].values:
                            print(f"📋 Înregistrarea pentru {data['Nume']} {data['Prenume']} există deja în Excel")
                            return True  # Considerăm că este ok, nu este o eroare
                            
                except:
                    # Dacă nu poate citi, creăm unul nou
                    df_existing = pd.DataFrame()
            else:
                # Creăm un DataFrame gol
                df_existing = pd.DataFrame()
            
            # Creăm DataFrame cu noua înregistrare
            df_new = pd.DataFrame([data])
            
            # Combinăm cu datele existente
            if not df_existing.empty:
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            else:
                df_combined = df_new
            
            # Ne asigurăm că CNP-ul și Telefonul sunt tratate ca string pentru a evita formatarea automată
            if 'CNP' in df_combined.columns:
                df_combined['CNP'] = df_combined['CNP'].astype(str)
            if 'Telefon' in df_combined.columns:
                df_combined['Telefon'] = df_combined['Telefon'].astype(str)
            
            # Reordonăm coloanele pentru a respecta ordinea cerută
            preferred_columns = [
                'Nume', 'Initiala_Tatalui', 'Prenume', 'CNP', 'Adresa', 'ANAF_Apartin', 'Telefon', 'Email', '2_Ani'
            ]
            
            # Adăugăm coloanele suplimentare la sfârșit pentru referință
            additional_columns = [
                'Nume_Fisier', 'Folder_Sursa', 'Cale_Fisier', 'Data_Procesare', 'Text_Complet'
            ]
            
            # Reordonăm coloanele existente
            existing_main_columns = [col for col in preferred_columns if col in df_combined.columns]
            existing_additional_columns = [col for col in additional_columns if col in df_combined.columns]
            other_columns = [col for col in df_combined.columns if col not in preferred_columns + additional_columns]
            
            final_columns = existing_main_columns + existing_additional_columns + other_columns
            
            df_combined = df_combined[final_columns]
            
            # Salvăm în Excel cu formatare
            with pd.ExcelWriter(self.excel_file_path, engine='openpyxl') as writer:
                df_combined.to_excel(writer, sheet_name='Date_Persoane', index=False)
                
                # Obținem worksheet-ul pentru formatare
                worksheet = writer.sheets['Date_Persoane']
                
                # Formatăm coloanele CNP și Telefon ca text pentru a evita formatarea automată
                from openpyxl.styles import NamedStyle
                from openpyxl.utils import get_column_letter
                
                # Găsim coloanele CNP și Telefon
                columns_to_format = {'CNP': None, 'Telefon': None}
                for idx, col_name in enumerate(df_combined.columns, 1):
                    if col_name in columns_to_format:
                        columns_to_format[col_name] = idx
                
                print(f"Debug: Coloane de formatat găsite în add_single_record: {columns_to_format}")
                
                # Formatăm coloanele găsite
                for col_name, col_index in columns_to_format.items():
                    if col_index:
                        col_letter = get_column_letter(col_index)
                        # Formatăm întreaga coloană ca text (inclusiv header-ul)
                        for row in range(1, len(df_combined) + 2):  # De la header (1) până la ultimul rând
                            cell = worksheet[f'{col_letter}{row}']
                            cell.number_format = '@'  # Format text
                            
                            # Pentru datele din rândurile de conținut (nu header)
                            if row > 1 and cell.value is not None:
                                # Convertim totul la string, indiferent de conținut
                                original_value = str(cell.value).strip()
                                
                                # Pentru CNP, dacă este în notație științifică, îl convertim
                                if col_name == 'CNP' and ('E+' in original_value or 'e+' in original_value or '.' in original_value):
                                    try:
                                        # Încercăm să convertim din notație științifică
                                        cell.value = str(int(float(original_value)))
                                    except:
                                        cell.value = original_value
                                # Pentru telefon, eliminăm .0 dacă există, dar păstrăm zero-urile de la început
                                elif col_name == 'Telefon' and original_value.endswith('.0'):
                                    # Eliminăm doar .0 de la sfârșit, păstrând zero-urile de la început
                                    cell.value = original_value[:-2]
                                else:
                                    # Pentru toate celelalte, îl păstrăm ca string
                                    cell.value = original_value
                
                # Ajustăm lățimea coloanelor
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Maxim 50 caractere
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"📋 Adăugată în Excel: {data['Nume']} {data['Prenume']}")
            return True
            
        except Exception as e:
            print(f"Eroare la adăugarea înregistrării în Excel: {e}")
            return False

def create_excel_summary(output_folder):
    """Funcție principală pentru crearea rezumatului Excel"""
    try:
        excel_manager = ExcelManager(output_folder)
        success = excel_manager.process_all_txt_files(output_folder)
        
        if success:
            return excel_manager.excel_file_path
        else:
            return None
            
    except Exception as e:
        print(f"Eroare la crearea rezumatului Excel: {e}")
        return None

def add_single_person_to_excel(output_folder, txt_file_path):
    """Funcție pentru a adăuga o singură persoană în Excel"""
    try:
        excel_manager = ExcelManager(output_folder)
        return excel_manager.add_single_record_to_excel(txt_file_path)
    except Exception as e:
        print(f"Eroare la adăugarea persoanei în Excel: {e}")
        return False
